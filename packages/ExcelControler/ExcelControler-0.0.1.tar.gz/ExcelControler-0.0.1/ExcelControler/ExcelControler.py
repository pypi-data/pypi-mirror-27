import openpyxl
import string
from openpyxl import load_workbook
from operator import is_not
from functools import partial


class Loader:

    """
    use excel_to_memory :
    from excel to
    [
    [row1, row1, row1],
    [row2, row2, row2],
    [row3, row3, row3],
    ........ ]


        """

    def load_excel(self, file, sheet=0):
        self.wb = load_workbook(file)
        self.sheet_name = self.wb.sheetnames

        self.ws = self.wb[self.sheet_name[sheet]]  # sheet selector
        self.row = self.ws.max_row

        return self.ws, self.row

    def find_object(self, file, object, columns=list(string.ascii_uppercase), sheet=0):
        self.object_position = []
        self.ws, self.max_row = self.load_excel(file)
        for column in columns:
            for row in range(1, self.max_row):
                try:
                    if self.ws["%s%s" % (column, row)].value == object:
                        self.object_position.append([row, column])
                except ValueError:
                    pass
        return self.object_position

    def letter_to_number(self, letter):
        self.letter_dict = {}
        self.number = 0
        for l in list(string.ascii_uppercase):
            self.letter_dict[l] = self.number
            self.number += 1
        self.convert_number = self.letter_dict[letter]
        return self.convert_number

    def fetch_row_data(self, worksheet, row):
        self.columns = list(string.ascii_uppercase)
        self.columns.append("AA")
        self.columns.append("AB")
        self.columns.append("AC")
        self.data = []
        for column in self.columns:
            info = worksheet["%s%s" % (column, str(row))].value
            self.data.append(info)
        return self.data

    def excel_to_memory(self, file, sheet=0):
        self.ws, self.row = self.load_excel(file, sheet)
        self.all_data = []
        self.process_data = []
        for i in range(1, self.row + 100):
            try:
                row_data = self.fetch_row_data(self.ws, i)
            except ValueError:
                pass
            # print(row_data)
            self.all_data.append(row_data)
        return self.all_data


class DataManager:
    def filter_none_data(self, all_data):
        self.process_data = []
        for data in all_data:

            filter_all_none_data = list(filter(partial(is_not, None), data))
            if filter_all_none_data != []:
                self.process_data.append(filter_all_none_data)
        return self.process_data

    def get_column_element(self, index, target_data):  # get elements in a columns (without duplicate), return a list
        self.target_list = []

        for data in target_data:
            self.target_list.append(data[index])

        self.target_list = list(set(self.target_list))
        return self.target_list


class Writer:
    """
    from
    [
    [row1, row1, row1],
    [row2, row2, row2],
    [row3, row3, row3],
    ........ ]

    to excel  (.xlsx file)

        """

    def memory_to_excel(self, filename, memory):
        print("Exporting to excel...")
        columns = list(string.ascii_uppercase)
        row_number = len(memory)

        column_number = 0
        for row in memory:
            if len(row) >= column_number:
                column_number = len(row)

        # column_number = len(memory[0])
        wb = openpyxl.Workbook()
        ws = wb.active
        for row, items in zip(range(1, row_number + 1), memory):
            for column, cell in zip(columns[0:column_number], items):
                ws["%s%s" % (column, row)] = cell
        wb.save(filename)
        print("Done")


if __name__ == "__main__":
    pass
