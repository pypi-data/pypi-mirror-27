#! python3
# Example to get all values based on column index
import openpyxl
from collections import defaultdict


# from openpyxl.cell import get_column_letter,column_index_from_string

def excel_get_dict(workbook, sheet='Sheet1', column_ref='column1'):
    """

    :param workbook:   .xlsx file to read
    :param sheet:      sheet name on the workbook
    :param column_ref: column to use as main key for the dict.
    :return: dictionary with the data
    """

    # Open the spreadsheet and get the latest dues status.
    wb = openpyxl.load_workbook(workbook)
    sheet = wb.get_sheet_by_name(sheet)
    # Define columns index by Name with value in row 1
    # Example: {'status': 'C', 'Notes': 'K'}
    columns = defaultdict(dict)
    # for 1 to last column

    # Define the columns: {'column1': 1, 'column2': 2}
    for i in range(1, sheet.max_column):
        # set dict columns.setdefault('value', 'column')
        ref_column_value = sheet.cell(row=1, column=i).value
        # column_index is A, B, C, etc.
        column_index = sheet.cell(row=1, column=i).column

        columns[ref_column_value] = column_index
    # Define sheet dictionary.

    '''
    Create list of servers, example:
     'HName': {'IP': '192.168.69.5',
                   'OS': 'Vmware',
                   'description': 'Vmware esx server',
                   'status': 'active'}}
    '''

    datadict = defaultdict(dict)
    # Read each row of the sheet.
    for row in range(2, sheet.max_row + 1):  # Skip the first row
        # read each column of each row.
        # k is key value, like: Name
        # v is value of the key, the column index like A
        for k, v in sorted(columns.items()):
            # Set dictionary for each server in nested dict
            # key server
            # Nested key names and values based on each column key Name
            # like datadict.setdefault('servername', {})['IP'] = sheet['A5'].value

            # column_name = sheet['A2'].value
            column_ref_coord = '{}{}'.format(columns.get(column_ref), str(row))
            ref_column_value = sheet[column_ref_coord].value

            column_nested = k

            column_nested_coord = '{}{}'.format(v, str(row))
            column_nested_value = sheet[column_nested_coord].value

            # datadict['c1']['column2'] = 'value'
            datadict[ref_column_value][column_nested] = column_nested_value

    return datadict

    # Example:
    # file = 'servers.xlsx'
    # servers = excel_get_dict(workbook, sheet, column_ref)
    # servers = excel_get_dict(file, 'Servers','Name')


def excel_get_list(workbook, sheet='Sheet1'):
    """

    :param workbook:   .xlsx file to read
    :param sheet:      sheet name on the workbook
    :return (list): list with the data
    """

    # Open the spreadsheet and get the latest dues status.
    wb = openpyxl.load_workbook(workbook)
    sheet = wb.get_sheet_by_name(sheet)
    # Define columns index by Name with value in row 1
    # Example: {'status': 'C', 'Notes': 'K'}
    columns = defaultdict(dict)
    # for 1 to last column

    # Define the columns: {'column1': 1, 'column2': 2}
    for i in range(1, sheet.max_column):
        # set dict columns.setdefault('value', 'column')
        ref_column_value = sheet.cell(row=1, column=i).value
        # column_index is A, B, C, etc.
        column_index = sheet.cell(row=1, column=i).column

        columns[ref_column_value] = column_index

    # Define sheet dictionary.

    datalist = []

    # Read each row of the sheet.
    for row in range(2, sheet.max_row + 1):  # Skip the first row
        # read each column of each row.
        # k is key value, like: Name
        # v is value of the key, the column index like A
        rowdict = {}

        for k, v in sorted(columns.items()):
            column = k

            column_coord = '{}{}'.format(v, str(row))
            column_value = sheet[column_coord].value

            # rowdict['column2'] = 'value'
            rowdict[column] = column_value

        datalist.append(rowdict)

    return datalist

    # Example:
    # file = 'servers.xlsx'
    # servers = excel_get_dict(workbook, sheet, column_ref)
    # servers = excel_get_dict(file, 'Servers','Name')
