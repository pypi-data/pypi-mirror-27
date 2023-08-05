# ExcelOpxWrapper : Can be run anywhere but only with post 2010 Excel formats
# noinspection PyBroadException
try:
    import numpy as np
except:
    pass

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from .excel_util import ERR_NA, ERR_VALUE, ERR_REF, ERR_NAME, ERR_DIV
# from openpyxl.cell import Cell


import os
from os import path

import abc
from abc import abstractmethod

class ExcelWrapper(object, metaclass=abc.ABCMeta):
    @abstractmethod
    def rangednames(self):
        return

    @abstractmethod
    def connect(self):
        return

    @abstractmethod
    def save(self):
        return

    @abstractmethod
    def save_as(self, filename, delete_existing=False):
        return

    @abstractmethod
    def close(self):
        return

    @abstractmethod
    def quit(self):
        return

    @abstractmethod
    def set_sheet(self, s):
        return

    @abstractmethod
    def get_sheet(self):
        return

    @abstractmethod
    def get_range(self, range):
        return

    @abstractmethod
    def get_used_range(self):
        return

    @abstractmethod
    def get_active_sheet(self):
        return

    @abstractmethod
    def get_cell(self, r, c):
        return

    def get_value(self, r, c):
        return self.get_cell(r, c).value

    def set_value(self, r, c, val):
        self.get_cell(r, c).value = val

    def get_formula(self, r, c):
        f = self.get_cell(r, c).Formula
        return f if f.startswith("=") else None

    def has_formula(self, range):
        f = self.get_range(range).Formula
        if type(f) == str:
            return f and f.startswith("=")
        else:
            for t in f:
                if t[0].startswith("="):
                    return True
            return False

    def get_formula_from_range(self, range):
        f = self.get_range(range).Formula
        if isinstance(f, (list, tuple)):
            if any([x for x in f if x[0].startswith("=")]):
                return [x[0] for x in f]
            else:
                return None
        else:
            return f if f.startswith("=") else None

    def get_formula_or_value(self, name):
        r = self.get_range(name)
        return r.Formula or r.Value

    @abstractmethod
    def get_row(self, row):
        """"""
        return

    @abstractmethod
    def set_calc_mode(self, automatic=True):
        """"""
        return

    @abstractmethod
    def set_screen_updating(self, update):
        """"""
        return

    @abstractmethod
    def run_macro(self, macro):
        """"""
        return
# Excel range wrapper that distribute reduced api used by compiler (Formula & Value)


class OpxRange(object):
    def __init__(self, cells, cells_do):

        super(OpxRange, self).__init__()

        self.cells = cells
        self.cellsDO = cells_do

        # Added for caching as the function Formula is used many times, and I don't think the cells content changes.
        self.formulas = None
        self.values = None

    def reset(self):
        self.formulas = None
        self.values = None

    @property
    def Formula(self):
        if not self.formulas:
            self.formulas = ()
            for row in self.cells:
                col = ()
                for cell in row:
                    col += (str(cell.value),)
                self.formulas += (col,)
            if sum(map(len, self.formulas)) == 1:
                self.formulas = self.formulas[0][0]

        return self.formulas

    @property
    def value(self):
        if not self.values:
            self.values = []
            for row in self.cellsDO:
                col = ()
                for cell in row:
                    if cell.data_type == 'f':
                        col += (cell.value,)
                    elif cell.data_type == 'n':
                        col += (cell.value,)
                    elif cell.data_type == 's':
                        col += (cell.value,)
                    elif cell.data_type == 'b':
                        col += (cell.value,)
                    # we assume it concerns a formula with an error as it's input isn't correct yet.
                    elif cell.data_type == 'e' and cell.value and cell.value == ERR_NA:
                        # It's better not to treat the #N/A value as None as it might result in unncessary evaluations
                        col += (ERR_NA,)
                    elif cell.data_type == 'e' and cell.value and cell.value == ERR_VALUE:
                        col += (ERR_VALUE,)
                    elif cell.data_type == 'e' and cell.value and cell.value == ERR_REF:
                        col += (ERR_REF,)
                    elif cell.data_type == 'e' and cell.value and cell.value == ERR_NAME:
                        col += (ERR_NAME,)
                    elif cell.data_type == 'e' and cell.value and cell.value == ERR_DIV:
                        col += (ERR_DIV,)
                    # We raise an exception as it might result in strange errors:
                    # It's highly possible the dimension of the self.values doesn't correspond to the dimension of
                    # that of the self.formulas such that we could match a formula with the wrong value
                    else:  # We always add a value, else the dimensions of the values
                        raise Exception(
                            "Unknown cell data type: [%s], value: [%s], cell: [%s]. Ensure to support it to overcome dimension mismatch with formulas" % (
                            cell.data_type, cell.value, cell))
                        # col += None

                self.values += (col,)
            if sum(map(len, self.values)) == 1:
                self.values = self.values[0][0]

        return self.values

# OpenPyXl implementation for ExcelWrapper interface
class ExcelOpxWrapper(ExcelWrapper):
    def __init__(self, filename, ):

        super(ExcelWrapper, self).__init__()

        self.filename = path.abspath(filename)

    @property
    def rangednames(self):

        if self.workbook is None:
            return None

        rangename = []
        for named_range in self.workbook.get_named_ranges():
            for worksheet, range_alias in named_range.destinations:
                tuple_name = (len(rangename) + 1, str(named_range.name), str(worksheet.title + '!' + range_alias))
                rangename.append([tuple_name])
        return rangename

    def connect(self):
        self.workbook = load_workbook(self.filename)
        self.workbookDO = load_workbook(self.filename, data_only=True)

    def save(self):
        self.workbook.save(self.filename)

    def save_as(self, filename, delete_existing=False):
        if delete_existing and os.path.exists(filename):
            os.remove(filename)
        self.workbook.save(filename)

    def close(self):
        return

    def quit(self):
        return

    def set_sheet(self, s):
        self.workbook.active = self.workbook.index(self.workbook[s])
        self.workbookDO.active = self.workbookDO.index(self.workbookDO[s])

        # Removed, as it's deprecated.
        # self.workbook.active = self.workbook.get_index(self.workbook[s])
        # self.workbookDO.active = self.workbookDO.get_index(self.workbookDO[s])
        return self.workbook.active

    def get_sheet(self):
        return self.workbook.active

    def max_rows(self):
        return self.workbook.active.max_row

    def max_columns(self):
        return self.workbook.active.max_column

    def get_range(self, address):
        sheet = self.workbook.active
        sheetDO = self.workbookDO.active
        if address.find('!') > 0:
            title, address = address.split('!')
            sheet = self.workbook[title]
            sheetDO = self.workbookDO[title]

        min_col, min_row, max_col, max_row = range_boundaries(address.upper())
        cells = [[cell for cell in row] for row in
                 sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)]
        cellsDO = [[cell for cell in row] for row in
                   sheetDO.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)]

        # Removed, as it's deprecated.
        # cells = [[cell for cell in row] for row in sheet.iter_rows(address)]
        # cellsDO = [[cell for cell in row] for row in sheetDO.iter_rows(address)]

        return OpxRange(cells, cellsDO)

    def get_used_range(self):
        return self.workbook.active.iter_rows()

    def get_active_sheet(self):
        return self.workbook.active.title

    def get_cell(self, r, c):
        # this could be improved in order not to call get_range
        return self.get_range(self.workbook.active.cell(None, r, c).coordinate)

    def get_row(self, row):
        return [self.get_value(row, col + 1) for col in range(self.workbook.active.max_column)]

    def set_calc_mode(self, automatic=True):
        raise Exception('Not implemented')

    def set_screen_updating(self, update):
        raise Exception('Not implemented')

    def run_macro(self, macro):
        raise Exception('Not implemented')
